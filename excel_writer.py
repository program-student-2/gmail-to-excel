# -*- coding: utf-8 -*-
"""抽出結果をExcel（サマーインターン管理表）へ転記するモジュール。"""

import re

import openpyxl

# 企業名の正規化で取り除く語句
_NOISE = ["株式会社", "（株）", "(株)", "グループ", "ホールディングス", "  ", " ", "　"]


def _normalize(name):
    if not name:
        return ""
    n = name.strip()
    for noise in _NOISE:
        n = n.replace(noise, "")
    return n.lower()


def _company_matches(extracted, existing):
    """企業名が同一企業を指すかどうかを判定する。"""
    a, b = _normalize(extracted), _normalize(existing)
    if not a or not b:
        return False
    if a == b:
        return True
    # 片方がもう片方を含む（3文字以上のときのみ。誤マッチ防止）
    shorter, longer = (a, b) if len(a) <= len(b) else (b, a)
    return len(shorter) >= 3 and shorter in longer


class ExcelTable:
    def __init__(self, path, sheet_name, columns, mypage_mark="済"):
        self.path = path
        self.wb = openpyxl.load_workbook(path)
        self.ws = self.wb[sheet_name]
        self.col = columns  # {"company":1, "mypage_registered":2, ...}
        self.mypage_mark = mypage_mark

    def company_names(self):
        """既存の企業名（空でないもの）の一覧を返す。"""
        c = self.col["company"]
        names = []
        for row in range(2, self.ws.max_row + 1):
            v = self.ws.cell(row=row, column=c).value
            if v not in (None, ""):
                names.append(str(v))
        return names

    def _find_row(self, company_name, existing_match=None):
        """企業名に一致する既存行番号を返す。無ければ None。

        existing_match（Claudeが返した既存表記）があれば最優先で完全一致を探す。
        """
        c = self.col["company"]
        if existing_match:
            for row in range(2, self.ws.max_row + 1):
                v = self.ws.cell(row=row, column=c).value
                if v is not None and str(v) == existing_match:
                    return row
        for row in range(2, self.ws.max_row + 1):
            existing = self.ws.cell(row=row, column=c).value
            if existing and _company_matches(company_name, str(existing)):
                return row
        return None

    def _first_empty_row(self):
        c = self.col["company"]
        row = 2
        while self.ws.cell(row=row, column=c).value not in (None, ""):
            row += 1
        return row

    def _set_if_allowed(self, row, col_key, value, overwrite, changes, label):
        """セルに値を入れる。空欄のときだけ（overwrite=Trueなら常に）更新する。"""
        if not value:
            return
        col = self.col[col_key]
        current = self.ws.cell(row=row, column=col).value
        if current in (None, "") or overwrite:
            if str(current or "") != str(value):
                self.ws.cell(row=row, column=col).value = value
                changes.append(f"{label}={value}")

    def upsert(self, data, overwrite=False):
        """抽出結果1件をExcelへ反映。(状態, 行番号, 変更内容) を返す。"""
        company = data.get("company_name", "").strip()
        if not company:
            return ("skip", None, ["企業名なし"])

        row = self._find_row(company, data.get("existing_match"))
        status = "update"
        if row is None:
            row = self._first_empty_row()
            self.ws.cell(row=row, column=self.col["company"]).value = company
            status = "new"

        changes = []
        self._set_if_allowed(row, "id", data.get("login_id"), overwrite, changes, "ID")
        self._set_if_allowed(row, "pw", data.get("password"), overwrite, changes, "PW")
        self._set_if_allowed(row, "url", data.get("login_url"), overwrite, changes, "URL")

        # マイページ登録欄に印を立てる（IDかURLが取れたとき）
        if data.get("login_id") or data.get("login_url"):
            mp = self.col["mypage_registered"]
            if self.ws.cell(row=row, column=mp).value in (None, ""):
                self.ws.cell(row=row, column=mp).value = self.mypage_mark
                changes.append(f"マイページ登録={self.mypage_mark}")

        if status == "new" and not changes:
            changes.append("新規行（企業名のみ）")
        return (status, row, changes)

    def save(self):
        self.wb.save(self.path)
